import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color
# Input/Output
INPUT_FILE = "NCS520 and UAR600D Feature Compare-1226.xlsx"
OUTPUT_FILE = "Formatted_Feature_Compare.xlsx"
def transform():
    print(f"Reading {INPUT_FILE}...")
    try:
        # Explicitly specify sheet name
        df = pd.read_excel(INPUT_FILE, sheet_name="UAR600D-10XA Feature")
    except Exception as e:
        print(f"Pandas read failed, trying openpyxl: {e}")
        df = pd.read_excel(INPUT_FILE, sheet_name="UAR600D-10XA Feature", engine='openpyxl')
    
    # Normalize columns
    print("Columns:", df.columns.tolist())
    
    # Ensure '描述' exists for grouping
    if '描述' not in df.columns:
        print("Error: Column '描述' not found.")
        return
        
    # Forward fill 描述 to get groups
    df['Group'] = df['描述'].ffill()
    
    # Create WB
    wb = Workbook()
    ws = wb.active
    ws.title = "Feature Compare"
    
    # Headers
    headers = ['描述', '规格', 'UT NOS', 'Comments']
    ws.append(headers)
    
    # Styling Constants
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    BOLD_FONT = Font(bold=True)
    RED_FONT = Font(color="FF0000")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Style Main Header
    for cell in ws[1]:
        cell.font = BOLD_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    current_row = 2
    current_group = None
    
    # Iterate groups
    for idx, row in df.iterrows():
        group = row['Group']
        
        # New Section?
        if group != current_group:
            # Add Section Header (A-D)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            cell = ws.cell(row=current_row, column=1)
            cell.value = group
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            for c in range(1, 5):
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            
            current_row += 1
            current_group = group
            
        # Data Row
        data = [
            "", # A
            row.get('规格', ""),
            row.get('UT NOS', ""),
            row.get('Comments', "")
        ]
        
        # Write row
        for c_idx, val in enumerate(data):
            c_num = c_idx + 1
            cell = ws.cell(row=current_row, column=c_num)
            if pd.isna(val):
                cell.value = ""
            else:
                cell.value = val
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN
            
            # Special alignment for UT NOS (Col C)
            if c_num == 3:
                cell.alignment = CENTER_ALIGN
                
            # Conditional Formatting: RED if UT NOS == 'N'
            ut_nos_val = str(row.get('UT NOS', "")).strip().upper()
            if ut_nos_val == 'N':
                cell.font = RED_FONT
                
        current_row += 1

    # Adjust Widths
    dims = {'A': 15, 'B': 50, 'C': 15, 'D': 40}
    for col, width in dims.items():
        ws.column_dimensions[col].width = width
    
    wb.save(OUTPUT_FILE)
    print(f"Done. Saved to {OUTPUT_FILE}")
if __name__ == "__main__":
    transform()