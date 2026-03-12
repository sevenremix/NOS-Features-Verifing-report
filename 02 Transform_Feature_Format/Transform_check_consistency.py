import pandas as pd
from openpyxl import load_workbook

INPUT_FILE = "NCS520 and UAR600D Feature Compare-1226.xlsx"
OUTPUT_FILE = "Formatted_Feature_Compare.xlsx"

def normalize(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() == 'nan':
        return ""
    return s

def get_input_data():
    # Read the specific sheet
    df = pd.read_excel(INPUT_FILE, sheet_name="UAR600D-10XA Feature")
    df['Group'] = df['描述'].ffill()
    data = []
    for _, row in df.iterrows():
        data.append({
            'Group': normalize(row['Group']),
            '规格': normalize(row['规格']),
            'UT_NOS': normalize(row.get('UT NOS', ""))
        })
    return data

def get_output_data():
    wb = load_workbook(OUTPUT_FILE, data_only=True)
    ws = wb.active
    data = []
    current_group = None
    
    for r in range(2, ws.max_row + 1):
        # A:描述/Group, B:规格, C:UT NOS, D:Comments
        val_a = normalize(ws.cell(row=r, column=1).value)
        val_b = normalize(ws.cell(row=r, column=2).value)
        val_c = normalize(ws.cell(row=r, column=3).value)
        val_d = normalize(ws.cell(row=r, column=4).value)
        
        # Check if it's a section header (merged A-D)
        if val_a != "" and val_b == "" and val_c == "" and val_d == "":
            current_group = val_a
            continue
            
        if current_group:
            # Data row
            data.append({
                'Group': current_group,
                '规格': val_b,
                'UT_NOS': val_c
            })
            
    return data

def compare():
    print(f"Comparing {INPUT_FILE} (UAR600D-10XA Feature) -> {OUTPUT_FILE}")
    input_list = get_input_data()
    output_list = get_output_data()
    
    # Convert to sets of tuples for easy comparison
    # (Group, 规格, UT_NOS)
    input_set = sorted([(d['Group'], d['规格'], d['UT_NOS']) for d in input_list])
    output_set = sorted([(d['Group'], d['规格'], d['UT_NOS']) for d in output_list])
    
    if len(input_set) != len(output_set):
        print(f"Warning: Row counts differ. Input: {len(input_set)}, Output: {len(output_set)}")
    
    mismatches = []
    from collections import Counter
    input_counts = Counter(input_set)
    output_counts = Counter(output_set)
    
    all_keys = set(input_counts.keys()) | set(output_counts.keys())
    
    for key in sorted(list(all_keys)):
        if input_counts[key] != output_counts[key]:
            mismatches.append((key, input_counts[key], output_counts[key]))
            
    if not mismatches:
        print("Success: All '规格' and 'UT NOS' data matches perfectly!")
    else:
        print(f"Failed: Found {len(mismatches)} discrepancies.")
        for item, in_c, out_c in mismatches[:10]:
            print(f"  Item: {item}")
            print(f"    Input Count: {in_c}, Output Count: {out_c}")
        if len(mismatches) > 10:
            print("  ...")

if __name__ == "__main__":
    compare()
