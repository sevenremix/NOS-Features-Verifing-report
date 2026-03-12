from openpyxl import load_workbook
import pandas as pd

def verify_report(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"--- Verifying {file_path} ---")
    
    # 1. Check Headers
    expected_headers = ["Feature Category", "Feature", "UT NOS", "Comments", "Case items", "Test Result", "Test Category", "CLI State", "Test Version"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
    print(f"Headers: {headers}")
    assert headers == expected_headers, f"Header mismatch! Expected {expected_headers}, got {headers}"
    
    # 2. Check Column Widths (F=6, H=8)
    # Note: openpyxl width might be slightly different than what was set (due to internal scaling)
    # We set 20.1
    width_f = ws.column_dimensions['F'].width
    width_h = ws.column_dimensions['H'].width
    print(f"Column F width: {width_f}, Column H width: {width_h}")
    assert 20.0 <= width_f <= 21.0, f"Column F width {width_f} not in range [20.0, 21.0]"
    assert 20.0 <= width_h <= 21.0, f"Column H width {width_h} not in range [20.0, 21.0]"
    
    # 3. Check Last Row
    last_row = ws.max_row
    print(f"Last row: {last_row}")
    
    # Check "Total Statistics:" label in merged cells A-D
    val_a = ws.cell(row=last_row, column=1).value
    print(f"Last row col 1 value: {val_a}")
    assert val_a == "Total Statistics:", f"Expected 'Total Statistics:', got {val_a}"
    
    # Check Stats Formatting (F and H)
    val_f = ws.cell(row=last_row, column=6).value
    val_h = ws.cell(row=last_row, column=8).value
    print(f"Last row col 6 value:\n{val_f}")
    print(f"Last row col 8 value:\n{val_h}")
    assert "\n" in str(val_f), "New-line missing in Test Result stats"
    assert "\n" in str(val_h), "New-line missing in CLI State stats"
    
    # Check Alignment of last row
    for c in range(1, 10):
        cell = ws.cell(row=last_row, column=c)
        align = cell.alignment
        print(f"Col {c} alignment: horiz={align.horizontal}, vert={align.vertical}, wrap={align.wrap_text}")
        assert align.horizontal == 'center', f"Col {c} not centered horizontally"
        assert align.vertical == 'center', f"Col {c} not centered vertically"
        assert align.wrap_text == True, f"Col {c} wrap_text should be True"

    # 4. Check Test Version handling
    # Sample some rows to ensure no 'nan'
    found_nan = False
    for r in range(2, last_row):
        val = ws.cell(row=r, column=9).value
        if val is not None and str(val).lower() == 'nan':
            found_nan = True
            print(f"Found 'nan' at row {r}, col 9")
            break
    assert not found_nan, "Found 'nan' in Test Version column"

    print("--- Verification Successful! ---")

if __name__ == "__main__":
    verify_report("e:\\NOS Feature & Case No.2\\Feature_TestCase (20260126).xlsx")
