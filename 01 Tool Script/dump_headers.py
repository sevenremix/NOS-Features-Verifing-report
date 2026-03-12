import pandas as pd
from openpyxl import load_workbook

TEST_FILE = "UAR600D-10XA Base Function Test Report (20260303).xlsx"

def dump_headers(file_path):
    print(f"Dumping headers for {file_path}...")
    wb = load_workbook(file_path, read_only=True)
    for sheet in wb.sheetnames:
        print(f"\nSheet: {sheet}")
        df = pd.read_excel(file_path, sheet_name=sheet, nrows=5)
        print(df.columns.tolist())
        # Print first valid data row
        print(df.iloc[0].tolist())

if __name__ == "__main__":
    dump_headers(TEST_FILE)
