import pandas as pd
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

file_path = "UAR600D-10XA Base Function Test Report.xlsx"

try:
    print("Trying default read...")
    df = pd.read_excel(file_path, sheet_name=None) # Read all sheets
    print("Sheets:", df.keys())
except Exception as e:
    print(f"Default read failed: {e}")
    try:
        print("Trying with engine='openpyxl' and no styles (if possible)...")
        # openpyxl doesn't have a direct 'no styles' flag in pandas read_excel
        # but we can try to use openpyxl directly to get values
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True, read_only=True)
        print("Sheets (openpyxl):", wb.sheetnames)
        for name in ["基础功能", "BFD", "IGMP", "拓展测试"]:
            if name in wb.sheetnames:
                ws = wb[name]
                cols = [cell.value for cell in next(ws.rows)]
                print(f"Columns for {name}:", cols)
    except Exception as e2:
        print(f"Openpyxl read failed: {e2}")
