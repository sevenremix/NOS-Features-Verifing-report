import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import warnings
import datetime
import glob

# Suppress warnings
warnings.filterwarnings("ignore")

DEFAULT_FEATURE_FILE = "Formatted_Feature_Compare.xlsx"

def read_xlsx_xml_robust(file_path, sheet_names=None):
    """
    Reads an xlsx file by parsing its XML directly, ignoring all styling.
    Useful for robust extraction from potentially complex or corrupted Excel files.
    """
    data = {}
    with zipfile.ZipFile(file_path, 'r') as z:
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            ss_xml = z.read('xl/sharedStrings.xml')
            root = ET.fromstring(ss_xml)
            for si in root.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                text_fragments = [t.text for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t') if t.text]
                shared_strings.append("".join(text_fragments))

        wb_xml = z.read('xl/workbook.xml')
        wb_root = ET.fromstring(wb_xml)
        sheet_map = {}
        for sheet in wb_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
            name = sheet.get('name')
            r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            sheet_map[name] = r_id

        rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_root = ET.fromstring(rels_xml)
        rel_map = {}
        for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
            rid = rel.get('Id')
            target = rel.get('Target')
            rel_map[rid] = target

        if sheet_names is None:
            sheet_names = sheet_map.keys()

        for name in sheet_names:
            if name not in sheet_map: continue
            r_id = sheet_map[name]
            target = rel_map[r_id]
            sheet_path = f"xl/{target}" if not target.startswith('xl/') else target
            if sheet_path not in z.namelist(): continue
            
            sheet_xml = z.read(sheet_path)
            s_root = ET.fromstring(sheet_xml)
            
            rows = []
            for row in s_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                row_data = {}
                for cell in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    r = cell.get('r')
                    t = cell.get('t')
                    v_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    if v_elem is not None:
                        val = v_elem.text
                        if t == 's': val = shared_strings[int(val)]
                        col_idx = "".join([c for c in r if c.isalpha()])
                        row_data[col_idx] = val
                rows.append(row_data)
            
            if rows:
                df = pd.DataFrame(rows)
                sorted_cols = sorted(df.columns, key=lambda x: (len(x), x))
                df = df[sorted_cols]
                header_idx = 0
                for i, r_dict in enumerate(rows):
                    if any(isinstance(v, str) and ("Feature" in v or "ID" in v) for v in r_dict.values()):
                        header_idx = i
                        break
                
                new_header = df.iloc[header_idx]
                df = df[header_idx+1:]
                df.columns = [str(c).strip() if pd.notna(c) else f"col_{i}" for i, c in enumerate(new_header)]
                data[name] = df.reset_index(drop=True)
    return data

def parse_sections_from_excel(file_path):
    """Load and parse sections from a local Excel file."""
    wb_feat = load_workbook(file_path, data_only=True)
    ws_feat = wb_feat.worksheets[0] # Use first sheet
    
    data_rows = []
    for r in range(2, ws_feat.max_row + 1):
        row = [ws_feat.cell(row=r, column=c).value for c in range(1, 5)]
        data_rows.append(row)
    return _convert_rows_to_sections(data_rows)

def parse_sections_from_list(raw_data):
    """Load and parse sections from a raw data list (e.g., from Feishu Sheet)."""
    # raw_data is expected to be a list of lists: [[col1, col2, col3, col4], ...]
    if not raw_data:
        return []
    # Skip header if it looks like one (e.g., first cell is "描述")
    start_idx = 1 if raw_data[0][0] == "描述" else 0
    return _convert_rows_to_sections(raw_data[start_idx:])

def _convert_rows_to_sections(rows):
    """Internal helper to convert raw row data into structured sections."""
    sections = []
    current_section = None
    
    for row in rows:
        # Expected row structure: [A:描述/Section, B:规格, C:UT_NOS, D:Comments]
        if len(row) < 2: continue
        a, b, c, d = row[0], row[1], row[2], row[3] if len(row) > 3 else None
        
        # Section Header (Merged A-D, B/C/D are empty/None)
        if a and not b and not c:
            sec_name = str(a).strip()
            if not sec_name.endswith("(一级Feature)"):
                sec_name += "(一级Feature)"
            current_section = {"name": sec_name, "features": []}
            sections.append(current_section)
        elif b and current_section:
            current_section["features"].append({
                "规格": str(b).strip(),
                "UT_NOS": str(c).strip() if c else "",
                "Comments": str(d).strip() if d else ""
            })
    return sections

def run_merge(feature_source, test_file, output_file=None):
    """
    Main programmatic interface for merging features and test results.
    feature_source: Can be a local file path (str) or a list of data rows (list).
    test_file: Local path to the test report Excel.
    output_file: Optional path for the generated report.
    """
    print(f"Loading Feature Data from {'file' if isinstance(feature_source, str) else 'list'}...")
    if isinstance(feature_source, str):
        sections = parse_sections_from_excel(feature_source)
    else:
        sections = parse_sections_from_list(feature_source)
        
    if not sections:
        print("Error: No feature sections found.")
        return None

    print(f"Loaded {len(sections)} sections.")

    print(f"Loading Test Report Details from {test_file}...")
    test_sheets = read_xlsx_xml_robust(test_file, ["基础功能", "BFD", "IGMP", "拓展测试"])
    
    all_cases = []
    for sheet_name, df in test_sheets.items():
        col_map = {}
        for c in df.columns:
            if "Item" in c: col_map["Item"] = c
            elif "Result" in c: col_map["Result"] = c
            elif "Feature" in c: col_map["Feature"] = c
            elif "CLI" in c: col_map["CLI"] = c
            elif "SW" in c: col_map["SW"] = c
        
        for _, row in df.iterrows():
            feat_raw = str(row.get(col_map.get("Feature"), "")).strip()
            if not feat_raw or feat_raw.lower() == 'nan': continue
            
            feat_list = [f.strip() for f in feat_raw.replace('\r\n', '\n').split('\n') if f.strip()]
            res_val = row.get(col_map.get("Result"), "")
            res_raw = str(res_val).strip().upper() if pd.notna(res_val) else ""
            
            # Map result
            if res_raw == 'P': res = 'Pass'
            elif res_raw == 'F': res = 'Fail'
            elif res_raw == 'N': res = 'Block'
            else: res = 'Incomplete'
            
            display_res = res_raw if res_raw else ""
            cli_val = row.get(col_map.get("CLI"), "")
            cli_raw = str(cli_val).strip() if pd.notna(cli_val) else ""
            
            for feat_name in feat_list:
                all_cases.append({
                    "Feature": feat_name,
                    "Case_Item": str(row.get(col_map.get("Item"), "")).strip(),
                    "Result": res,
                    "DisplayResult": display_res,
                    "Version": str(row.get(col_map.get("SW"), "")).strip(),
                    "CLI": cli_raw,
                    "Category": sheet_name
                })

    print(f"Loaded {len(all_cases)} test cases.")

    print("Generating Merged Report...")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Feature_TestCase"
    
    headers = ["描述（一级Feature）", "规格", "对应UT Feature", "Test Remarks", "Case items", "Test Result", "Test Category", "CLI State", "Version"]
    ws_out.append(headers)
    
    # Styles
    BLUE_FILL = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    BOLD_FONT = Font(bold=True)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for c in range(1, 10):
        ws_out.cell(row=1, column=c).font = BOLD_FONT
        ws_out.cell(row=1, column=c).alignment = CENTER_ALIGN
        ws_out.cell(row=1, column=c).border = THIN_BORDER
    
    ws_out.freeze_panes = "A2"
    ws_out.sheet_properties.outlinePr.summaryBelow = False
    ws_out.sheet_properties.outlinePr.applyStyles = True
    ws_out.sheet_properties.outlinePr.showOutlineSymbols = True

    current_row = 2
    total_feat_100_test_pass = 0
    total_feat_100_cli_pass = 0
    total_all_cases = 0
    total_features_count = 0
    
    for sec in sections:
        sec_features = sec["features"]
        if not sec_features: continue
        
        sec_row_idx = current_row
        current_row += 1
        ws_out.row_dimensions[sec_row_idx].outlineLevel = 0
        
        total_sec_cases = 0
        sec_features_total = 0
        sec_features_passed_test = 0
        sec_features_passed_cli = 0
        
        for feat in sec_features:
            spec = feat["规格"]
            matched = [c for c in all_cases if c["Feature"] == spec]
            
            # Special case for ZTP
            if spec == "Zero touch provisioning (ZTP)*":
                matched.append({
                    "Feature": spec, "Case_Item": "", "Result": "Incomplete", 
                    "DisplayResult": "", "Version": "", "CLI": "", "Category": "Special"
                })
            
            total_features_count += 1
            ws_out.row_dimensions[current_row].outlineLevel = 1
            # Auto-unfold Layer 2/3
            is_hidden = False if ("Layer 2" in sec["name"] or "Layer 3" in sec["name"]) else True
            ws_out.row_dimensions[current_row].hidden = is_hidden 
            
            ws_out.cell(row=current_row, column=2).value = spec
            ws_out.cell(row=current_row, column=3).value = feat["UT_NOS"]
            ws_out.cell(row=current_row, column=4).value = feat["Comments"]
            
            case_count = len(matched)
            ws_out.cell(row=current_row, column=5).value = f"Cases count: {case_count}"
            
            feat_verified = sum(1 for c in matched if c["CLI"] == 'Y')
            ws_out.cell(row=current_row, column=8).value = f"Pass {(feat_verified/case_count*100):.0f}%" if case_count > 0 else "Pass 0%"
            
            pass_c = sum(1 for c in matched if c["Result"] == 'Pass')
            ws_out.cell(row=current_row, column=6).value = f"Pass {(pass_c/case_count*100):.0f}%" if case_count > 0 else "Pass 0%"
            
            if case_count > 0 and pass_c == case_count:
                total_feat_100_test_pass += 1
                sec_features_passed_test += 1
            if case_count > 0 and feat_verified == case_count:
                total_feat_100_cli_pass += 1
                sec_features_passed_cli += 1
            
            sec_features_total += 1
            for c in range(1, 10):
                cell = ws_out.cell(row=current_row, column=c)
                cell.fill = LIGHT_BLUE_FILL
                cell.border = THIN_BORDER
                is_centered = c in [3, 6, 7, 8, 9]
                cell.alignment = Alignment(horizontal='center' if is_centered else 'left', vertical='center', wrap_text=(c==6))
            
            current_row += 1
            for mc in matched:
                ws_out.row_dimensions[current_row].outlineLevel = 2
                ws_out.row_dimensions[current_row].hidden = is_hidden
                ws_out.cell(row=current_row, column=5).value = mc["Case_Item"]
                ws_out.cell(row=current_row, column=6).value = mc["DisplayResult"]
                ws_out.cell(row=current_row, column=7).value = mc["Category"]
                ws_out.cell(row=current_row, column=8).value = mc["CLI"]
                ws_out.cell(row=current_row, column=9).value = mc["Version"] if mc["Version"] != "nan" else ""
                
                for c in range(1, 10):
                    cell = ws_out.cell(row=current_row, column=c)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(horizontal='center' if c in [3,6,7,8,9] else 'left', vertical='center')
                current_row += 1
            
            total_sec_cases += case_count
            total_all_cases += case_count

        # Section Header
        ws_out.merge_cells(start_row=sec_row_idx, start_column=1, end_row=sec_row_idx, end_column=4)
        ws_out.cell(row=sec_row_idx, column=1).value = f"{sec['name']}\nFeature Num: {sec_features_total}"
        ws_out.row_dimensions[sec_row_idx].height = 35
        ws_out.cell(row=sec_row_idx, column=1).font = BOLD_FONT
        ws_out.cell(row=sec_row_idx, column=5).value = f"Cases count: {total_sec_cases}"
        
        if sec_features_total > 0:
            ws_out.cell(row=sec_row_idx, column=6).value = f"feature pass {(sec_features_passed_test/sec_features_total*100):.0f}%"
            ws_out.cell(row=sec_row_idx, column=8).value = f"feature pass {(sec_features_passed_cli/sec_features_total*100):.0f}%"
        
        for c in range(1, 10):
            cell = ws_out.cell(row=sec_row_idx, column=c)
            cell.fill = BLUE_FILL
            cell.border = THIN_BORDER
            cell.font = BOLD_FONT
            cell.alignment = Alignment(horizontal='center' if c in [1, 5, 6, 7, 8, 9] else 'left', vertical='center', wrap_text=(c in [1, 6]))

    # Global Summary
    p_test_feat = (total_feat_100_test_pass / total_features_count * 100) if total_features_count > 0 else 0
    p_cli_feat = (total_feat_100_cli_pass / total_features_count * 100) if total_features_count > 0 else 0
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws_out.cell(row=current_row, column=1).value = f"Total Statistics:\nFeature Num: {total_features_count}"
    ws_out.cell(row=current_row, column=5).value = f"Total cases count: {total_all_cases}"
    ws_out.cell(row=current_row, column=6).value = f"Feature Pass Rate {p_test_feat:.0f}%\nFeature Pass Num {total_feat_100_test_pass}"
    ws_out.cell(row=current_row, column=8).value = f"Feature Pass Rate {p_cli_feat:.0f}%\nFeature Pass Num {total_feat_100_cli_pass}"
    
    for c in range(1, 10):
        cell = ws_out.cell(row=current_row, column=c)
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_out.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

    # Column Widths
    dims = {'A': 15, 'B': 40, 'C': 10, 'D': 20, 'E': 60, 'F': 25.2, 'G': 15, 'H': 25.2, 'I': 20}
    for col, width in dims.items():
        ws_out.column_dimensions[col].width = width
    ws_out.column_dimensions['G'].hidden = True
    
    if not output_file:
        now_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"Feature_TestCase_report_{now_str}.xlsx"
        
    wb_out.save(output_file)
    print(f"Done. Saved to {output_file}")
    return output_file

def main():
    # Find latest test report file
    test_files = glob.glob("UAR600D-10XA Base Function Test Report (*).xlsx")
    test_file = sorted(test_files)[-1] if test_files else "UAR600D-10XA Base Function Test Report.xlsx"
    
    if not os.path.exists(DEFAULT_FEATURE_FILE):
        print(f"Error: {DEFAULT_FEATURE_FILE} not found.")
        return

    run_merge(DEFAULT_FEATURE_FILE, test_file)

if __name__ == "__main__":
    main()
