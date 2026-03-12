import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import warnings

# Suppress warnings
warnings.filterwarnings("ignore")

FEATURE_FILE = "Formatted_Feature_Compare.xlsx"
TEST_FILE = "UAR600D-10XA Base Function Test Report.xlsx"
OUTPUT_FILE = "Feature_TestCase.xlsx"

def read_xlsx_xml_robust(file_path, sheet_names=None):
    """
    Reads an xlsx file by parsing its XML directly, ignoring all styling.
    """
    data = {}
    with zipfile.ZipFile(file_path, 'r') as z:
        shared_strings = []
        if 'xl/sharedStrings.xml' in z.namelist():
            ss_xml = z.read('xl/sharedStrings.xml')
            root = ET.fromstring(ss_xml)
            # A shared string entry <si> can contain multiple <t> tags if it's rich text.
            # We must join them to get the full string for that index.
            for si in root.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si'):
                text_fragments = [t.text for t in si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t') if t.text]
                shared_strings.append("".join(text_fragments))

        wb_xml = z.read('xl/workbook.xml')
        wb_root = ET.fromstring(wb_xml)
        sheet_map = {}
        for sheet in wb_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
            name = sheet.get('name')
            r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
            sheet_map[name] = r_id

        rels_xml = z.read('xl/_rels/workbook.xml.rels')
        rels_root = ET.fromstring(rels_xml)
        rel_map = {}
        for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
            id = rel.get('Id')
            target = rel.get('Target')
            rel_map[id] = target

        if sheet_names is None:
            sheet_names = sheet_map.keys()

        for name in sheet_names:
            if name not in sheet_map: continue
            r_id = sheet_map[name]
            target = rel_map[r_id]
            sheet_path = f"xl/{target}" if not target.startswith('xl/') else target
            if sheet_path not in z.namelist(): continue
            
            sheet_xml = z.read(sheet_path)
            s_root = ET.fromstring(sheet_xml)
            
            rows = []
            for row in s_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
                row_data = {}
                for cell in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                    r = cell.get('r')
                    t = cell.get('t')
                    v_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    if v_elem is not None:
                        val = v_elem.text
                        if t == 's': val = shared_strings[int(val)]
                        col_idx = "".join([c for c in r if c.isalpha()])
                        row_data[col_idx] = val
                rows.append(row_data)
            
            if rows:
                df = pd.DataFrame(rows)
                # Sort columns A, B, C...
                sorted_cols = sorted(df.columns, key=lambda x: (len(x), x))
                df = df[sorted_cols]
                # Find header row with 'Feature' or '\nID'
                header_idx = 0
                for i, r_dict in enumerate(rows):
                    if any(isinstance(v, str) and ("Feature" in v or "ID" in v) for v in r_dict.values()):
                        header_idx = i
                        break
                
                new_header = df.iloc[header_idx]
                df = df[header_idx+1:]
                df.columns = [str(c).strip() if pd.notna(c) else f"col_{i}" for i, c in enumerate(new_header)]
                data[name] = df.reset_index(drop=True)

    return data

def main():
    print("Loading Feature List...")
    # Read formatted feature file (using data_only to avoid formulas)
    wb_feat = load_workbook(FEATURE_FILE, data_only=True)
    ws_feat = wb_feat.active
    
    sections = []
    current_section = None
    
    for r in range(2, ws_feat.max_row + 1):
        a = ws_feat.cell(row=r, column=1).value
        b = ws_feat.cell(row=r, column=2).value
        c = ws_feat.cell(row=r, column=3).value
        d = ws_feat.cell(row=r, column=4).value
        
        # Section Header (Merged A-D, B/C/D are empty)
        if a and not b and not c and not d:
            sec_name = str(a).strip()
            if not sec_name.endswith("(一级Feature)"):
                sec_name += "(一级Feature)"
            current_section = {"name": sec_name, "features": []}
            sections.append(current_section)
        elif b and current_section:
            current_section["features"].append({
                "规格": str(b).strip(),
                "UT_NOS": str(c).strip() if c else "",
                "Comments": str(d).strip() if d else ""
            })

    print(f"Loaded {len(sections)} sections.")

    print("Loading Test Report Details...")
    test_sheets = read_xlsx_xml_robust(TEST_FILE, ["基础功能", "BFD", "IGMP", "拓展测试"])
    
    # Flatten test cases
    all_cases = []
    for sheet_name, df in test_sheets.items():
        # Identify columns
        # Expected: Item, Result, SW (Version), CLI, Feature
        col_map = {}
        for c in df.columns:
            if "Item" in c: col_map["Item"] = c
            elif "Result" in c: col_map["Result"] = c
            elif "Feature" in c: col_map["Feature"] = c
            elif "CLI" in c: col_map["CLI"] = c
            elif "SW" in c: col_map["SW"] = c
        
        for _, row in df.iterrows():
            feat_raw = str(row.get(col_map.get("Feature"), "")).strip()
            if not feat_raw or feat_raw.lower() == 'nan': continue
            
            # Feature column might contain multiple features separated by newlines
            feat_list = [f.strip() for f in feat_raw.replace('\r\n', '\n').split('\n') if f.strip()]
            
            res_val = row.get(col_map.get("Result"), "")
            res_raw = str(res_val).strip().upper() if pd.notna(res_val) else ""
            
            # Internal mapping for Stats (treat empty as Incomplete)
            if res_raw == 'P': res = 'Pass'
            elif res_raw == 'F': res = 'Fail'
            elif res_raw == 'N': res = 'Block'
            else: res = 'Incomplete'
            
            # Display mapping (preserve empty)
            display_res = res_raw if res_raw else ""
            
            cli_val = row.get(col_map.get("CLI"), "")
            cli_raw = str(cli_val).strip() if pd.notna(cli_val) else ""
            
            for feat_name in feat_list:
                all_cases.append({
                    "Feature": feat_name,
                    "Case_Item": str(row.get(col_map.get("Item"), "")).strip(),
                    "Result": res,
                    "DisplayResult": display_res,
                    "Version": str(row.get(col_map.get("SW"), "")).strip(),
                    "CLI": cli_raw,
                    "Category": sheet_name
                })

    print(f"Loaded {len(all_cases)} test cases.")

    # Create Output
    print("Generating Merged Report...")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Feature_TestCase"
    
    headers = ["Feature Category", "Feature", "UT NOS", "Comments", "Case items", "Test Result", "Test Category", "CLI State", "Test Version"]
    ws_out.append(headers)
    
    # Styles
    BLUE_FILL = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    LIGHT_BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    BOLD_FONT = Font(bold=True)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
    
    for c in range(1, 10):
        ws_out.cell(row=1, column=c).font = BOLD_FONT
        ws_out.cell(row=1, column=c).alignment = CENTER_ALIGN
        ws_out.cell(row=1, column=c).border = THIN_BORDER
    # Freeze Panes and Set Grouping Property
    ws_out.freeze_panes = "A2"
    ws_out.sheet_properties.outlinePr.summaryBelow = False

    current_row = 2
    
    # Global Counters
    total_feat_100_test_pass = 0
    total_feat_100_cli_pass = 0
    total_all_cases = 0
    total_features_count = 0
    
    for sec in sections:
        # Section Header calculation
        sec_start_row = current_row
        
        # Row for Level 1 Section Header
        sec_features = sec["features"]
        if not sec_features: continue
        
        sec_row_idx = current_row
        current_row += 1
        
        # Section Header is Level 0 (always visible)
        ws_out.row_dimensions[sec_row_idx].outlineLevel = 0
        
        total_sec_cases = 0
        sec_res_counts = {"Pass": 0, "Fail": 0, "Block": 0, "Incomplete": 0}
        total_sec_verified = 0
        
        for feat in sec_features:
            spec = feat["规格"]
            # Find matching cases
            matched = [c for c in all_cases if c["Feature"] == spec]
            
            # Special case for ZTP: Always add an empty placeholder as requested
            if spec == "Zero touch provisioning (ZTP)*":
                matched.append({
                    "Feature": spec,
                    "Case_Item": "", # Empty Case Item
                    "Result": "Incomplete", # Treat as incomplete for stats
                    "DisplayResult": "",
                    "Version": "",
                    "CLI": "",
                    "Category": "Special"
                })
            
            feat_start_row = current_row
            total_features_count += 1
            
            # Feature Summary row is Level 1 (collapsible under Section)
            ws_out.row_dimensions[current_row].outlineLevel = 1
            # Default hide except for Layer 3
            is_hidden = False if "Layer 3" in sec["name"] else True
            ws_out.row_dimensions[current_row].hidden = is_hidden 
            
            # Write row values
            ws_out.cell(row=current_row, column=1).value = "" # 描绘
            ws_out.cell(row=current_row, column=2).value = spec
            ws_out.cell(row=current_row, column=3).value = feat["UT_NOS"]
            ws_out.cell(row=current_row, column=4).value = feat["Comments"]
            
            case_count = len(matched)
            ws_out.cell(row=current_row, column=5).value = f"Cases count: {case_count}"
            
            # CLI Pass rate for feature
            feat_verified = sum(1 for c in matched if c["CLI"] == 'Y')
            if case_count > 0:
                fv_p = (feat_verified / case_count) * 100
                ws_out.cell(row=current_row, column=8).value = f"Pass {fv_p:.0f}%"
            else:
                ws_out.cell(row=current_row, column=8).value = "Pass 0%"
            
            # Stats for feature
            pass_c = sum(1 for c in matched if c["Result"] == 'Pass')
            fail_c = sum(1 for c in matched if c["Result"] == 'Fail')
            block_c = sum(1 for c in matched if c["Result"] == 'Block')
            inc_c = sum(1 for c in matched if c["Result"] == 'Incomplete')
            
            if case_count > 0:
                p_p = (pass_c / case_count) * 100
                stat_str = f"Pass {p_p:.0f}%"
            else:
                stat_str = "Pass 0%"

            # Check for 100% Pass for Global Stats
            is_test_100_pass = (case_count > 0 and pass_c == case_count)
            is_cli_100_pass = (case_count > 0 and feat_verified == case_count)
            
            if is_test_100_pass: total_feat_100_test_pass += 1
            if is_cli_100_pass: total_feat_100_cli_pass += 1
                
            cell_stat = ws_out.cell(row=current_row, column=6)
            cell_stat.value = stat_str
            cell_stat.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply Light Blue, borders, and Centering to feature summary
            for c in range(1, 10):
                cell = ws_out.cell(row=current_row, column=c)
                cell.fill = LIGHT_BLUE_FILL
                cell.border = THIN_BORDER
                # Alignment: All cells vertically centered
                is_centered_horiz = c in [3, 6, 7, 8, 9]
                cell.alignment = Alignment(horizontal='center' if is_centered_horiz else 'left', 
                                           vertical='center', 
                                           wrap_text=(c==6))
            
            current_row += 1
            
            # Detail rows are Level 2 (collapsible under Feature)
            for mc in matched:
                ws_out.row_dimensions[current_row].outlineLevel = 2
                # Default hide except for Layer 3
                is_hidden = False if "Layer 3" in sec["name"] else True
                ws_out.row_dimensions[current_row].hidden = is_hidden
                
                ws_out.cell(row=current_row, column=5).value = mc["Case_Item"]
                
                # Show empty if source was empty
                res_display = mc["DisplayResult"] if mc["DisplayResult"] else ""
                ws_out.cell(row=current_row, column=6).value = res_display
                
                ws_out.cell(row=current_row, column=7).value = mc["Category"]
                ws_out.cell(row=current_row, column=8).value = mc["CLI"] if mc["CLI"] else ""
                
                test_ver = str(mc["Version"]).strip() if mc["Version"] else ""
                if test_ver.lower() == 'nan': test_ver = ""
                ws_out.cell(row=current_row, column=9).value = test_ver
                
                for c in range(1, 10):
                    cell = ws_out.cell(row=current_row, column=c)
                    cell.border = THIN_BORDER
                    is_centered_horiz = c in [3, 6, 7, 8, 9]
                    cell.alignment = Alignment(horizontal='center' if is_centered_horiz else 'left', 
                                               vertical='center')
                current_row += 1
                
            # [REMOVED] Vertical Merge Feature info (User requested no merging for 描述 and 规格)
            # if current_row - 1 > feat_start_row:
            #     ws_out.merge_cells(start_row=feat_start_row, start_column=1, end_row=current_row - 1, end_column=1)
            #     ws_out.merge_cells(start_row=feat_start_row, start_column=2, end_row=current_row - 1, end_column=2)
            #     ws_out.merge_cells(start_row=feat_start_row, start_column=3, end_row=current_row - 1, end_column=3)
            #     ws_out.merge_cells(start_row=feat_start_row, start_column=4, end_row=current_row - 1, end_column=4)

            # Update section stats
            total_sec_cases += case_count
            total_all_cases += case_count
            sec_res_counts["Pass"] += pass_c
            sec_res_counts["Fail"] += fail_c
            sec_res_counts["Block"] += block_c
            sec_res_counts["Incomplete"] += inc_c
            total_sec_verified += feat_verified

        # Fill Section Header Row
        ws_out.merge_cells(start_row=sec_row_idx, start_column=1, end_row=sec_row_idx, end_column=4)
        ws_out.cell(row=sec_row_idx, column=1).value = sec["name"]
        ws_out.cell(row=sec_row_idx, column=1).font = BOLD_FONT
        
        ws_out.cell(row=sec_row_idx, column=5).value = f"Cases count: {total_sec_cases}"
        
        # CLI Pass rate for section
        if total_sec_cases > 0:
            sv_p = (total_sec_verified / total_sec_cases) * 100
            ws_out.cell(row=sec_row_idx, column=8).value = f"Pass {sv_p:.0f}%"
        else:
            ws_out.cell(row=sec_row_idx, column=8).value = "Pass 0%"
        
        if total_sec_cases > 0:
            sp_p = (sec_res_counts["Pass"] / total_sec_cases) * 100
            s_stat_str = f"Pass {sp_p:.0f}%"
        else:
            s_stat_str = "Pass 0%"
        
        cell_sec_stat = ws_out.cell(row=sec_row_idx, column=6)
        cell_sec_stat.value = s_stat_str
        
        for c in range(1, 10):
            cell = ws_out.cell(row=sec_row_idx, column=c)
            cell.fill = BLUE_FILL
            cell.border = THIN_BORDER
            cell.font = BOLD_FONT
            is_centered_horiz = c in [1, 5, 6, 7, 8, 9] # Section header cols
            cell.alignment = Alignment(horizontal='center' if is_centered_horiz else 'left', 
                                       vertical='center', 
                                       wrap_text=(c==6))

    # --- Global Summary Row ---
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Calculate Global Pass % for Features
    p_test_feat = (total_feat_100_test_pass / total_features_count * 100) if total_features_count > 0 else 0
    p_cli_feat = (total_feat_100_cli_pass / total_features_count * 100) if total_features_count > 0 else 0
        
    ws_out.cell(row=current_row, column=1).value = f"Total Statistics:\nFeature Num: {total_features_count}"
    ws_out.cell(row=current_row, column=5).value = f"Total cases count: {total_all_cases}"
    ws_out.cell(row=current_row, column=6).value = f"Feature Pass Rate {p_test_feat:.0f}%\nFeature Pass Num {total_feat_100_test_pass}"
    ws_out.cell(row=current_row, column=8).value = f"Feature Pass Rate {p_cli_feat:.0f}%\nFeature Pass Num {total_feat_100_cli_pass}"

    for c in range(1, 10):
        cell = ws_out.cell(row=current_row, column=c)
        cell.fill = YELLOW_FILL
        cell.border = THIN_BORDER
        cell.font = BOLD_FONT
        # Alignment: All cells centered and wrap text for the stats with newlines
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Merge after styling to ensure all cells have borders/fills
    ws_out.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)

    # Final logic: Adjust widths
    # Increased F width to 50 for stats
    # F (Test Result) and H (CLI State) set to 210 pixels (~25.2 in openpyxl units)
    dims = {'A': 15, 'B': 40, 'C': 10, 'D': 20, 'E': 60, 'F': 25.2, 'G': 15, 'H': 25.2, 'I': 20}
    for col, width in dims.items():
        ws_out.column_dimensions[col].width = width
    
    # Default hide Column G (Test category)
    ws_out.column_dimensions['G'].hidden = True
    
    # Freeze Panes
    ws_out.freeze_panes = "A2"
    
    wb_out.save(OUTPUT_FILE)
    print(f"Done. Saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
